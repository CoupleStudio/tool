#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import json
import os

src = os.path.join(os.getcwd(), "excel")
dst = os.path.join(os.getcwd(), "json")
if not os.path.exists(dst):
    os.makedirs(dst)

fileList = os.listdir(src)

for file in fileList:
    wb = load_workbook(os.path.join(src, file))
    ws = wb.active
    startRowIndex = 4
    startColIndex = 2
    key_id = 1
    key_name_row_index = 2
    dic = {}
    dic_keys = []
    for row in range(startRowIndex, ws.max_row+1):
        dic[str(ws.cell(row, key_id).value)] = {}
        for col in range(startColIndex, ws.max_column+1):
            dic[str(ws.cell(row, key_id).value)][str(
                ws.cell(key_name_row_index, col).value)] = ws.cell(row, col).value

    json_data = json.dumps(dic, ensure_ascii=False, indent=4, sort_keys=True)
    dstFile = file.split('.')[0] + ".json"
    newFile = os.path.join(os.getcwd(), "json", dstFile)
    fo = open(newFile, "w+")
    fo.write(json_data)
    fo.close()
